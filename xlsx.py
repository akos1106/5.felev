import openpyxl
from openpyxl.workbook import Workbook


from data.basic.model_dataclasses import Person
from data.basic.generator import generate_people


def write_peolpe(people: list[Person], wb: Workbook,
                 heading:bool=True,sheet_name:str="people") -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_name=["id","name","age","male"]
        for col in range(len(column_name)):
            sheet.cell(row=1,column=col+1,value=column_name[col])
    offset=2 if heading else 1

    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)


def read_people(wb:Workbook,sheet_name:str="people",
                heading:bool=True)->list[Person]:
    sheet=wb[sheet_name]

    people=[]
    row= 2 if heading else 1
    while True:
        cell=sheet.cell(row=row,column=1)
        if cell.value is None:
            break
        people.append(
            Person(
                sheet.cell(row=row,column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row+=1
    return people

if __name__ == "__main__":
    wb = Workbook()
    write_peolpe(
        generate_people(10),
        wb
    )
    wb.save("C:/Users/hallgato/people.xlsx")

    wb =openpyxl.load_workbook("C:/Users/hallgato/people.xlsx")
    for person in read_people(wb):
        print(person)